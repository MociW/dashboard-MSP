from openpyxl import Workbook
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def create_excel_base():
    """Create base workbook with default sheet removed."""
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    return wb


def create_styled_header(ws, value, start_row, start_col, end_row, end_col):
    """Create and style a header cell with optional merging."""
    cell = ws.cell(row=start_row, column=start_col, value=value)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    if start_row != end_row or start_col != end_col:
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=end_row,
            end_column=end_col,
        )


def apply_borders(ws, start_row, start_col, end_row, end_col):
    """Apply thin borders to a range of cells."""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
        for cell in row:
            cell.border = thin_border


def apply_threshold_formatting(cell, value, bounderies):
    """Apply conditional formatting based on value thresholds."""
    if not isinstance(value, (int, float)):
        return

    if value > bounderies:
        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    elif value < -bounderies:
        cell.fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")


def convert_to_excel_in_house(df, var_previous, var_current, bounderies):
    wb = create_excel_base()
    ws = wb.create_sheet(title="In House")

    # Populate data rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Apply threshold formatting to Gap columns
            if c_idx in [5, 8, 11, 14, 17] and isinstance(value, (int, float)):
                apply_threshold_formatting(cell, value, bounderies)

    # Create headers
    gap_columns = [(3, 5), (6, 8), (9, 11), (12, 14), (15, 17)]
    header_labels = ["LVA", "Non LVA", "Tooling", "Process Cost", "Total Cost"]

    create_styled_header(ws, "Part No", 1, 1, 2, 1)
    create_styled_header(ws, "Part Name", 1, 2, 2, 2)

    for i, (label, (start_col, end_col)) in enumerate(zip(header_labels, gap_columns)):
        create_styled_header(ws, label, 1, start_col, 1, end_col)
        create_styled_header(ws, f"{var_previous}", 2, start_col, 2, start_col)
        create_styled_header(ws, f"{var_current}", 2, start_col + 1, 2, start_col + 1)
        create_styled_header(ws, "Gap (%)", 2, end_col, 2, end_col)

    # Apply borders
    apply_borders(ws, 1, 1, 1, 17)
    apply_borders(ws, 2, 1, len(df) + 2, 17)

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def convert_to_excel_format_out_house(df, var_previous, var_current, bounderies):
    wb = create_excel_base()

    # Process each unique source
    for source in df["source"].unique():
        source_df = df[df["source"] == source]
        ws = wb.create_sheet(title=str(source))

        # Populate data
        for r_idx, row in enumerate(dataframe_to_rows(source_df, index=False, header=False), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Apply conditional formatting to Gap column
                if c_idx == 6:
                    apply_threshold_formatting(cell, value, bounderies)

        # Create headers
        create_styled_header(ws, "Part No", 1, 1, 2, 1)
        create_styled_header(ws, "Part Name", 1, 2, 2, 2)
        create_styled_header(ws, "Source", 1, 3, 2, 3)
        create_styled_header(ws, "Price", 1, 4, 1, 6)
        create_styled_header(ws, f"{var_previous}", 2, 4, 2, 4)
        create_styled_header(ws, f"{var_current}", 2, 5, 2, 5)
        create_styled_header(ws, "Gap (%)", 2, 6, 2, 6)

        # Apply borders
        apply_borders(ws, 1, 1, 1, 6)
        apply_borders(ws, 2, 1, len(source_df) + 2, 6)

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def convert_to_excel_format_out_house_per_part(df, var_previous, var_current, bounderies):
    df = df[
        [
            "part_num",
            "part_no",
            "part_name",
            "source",
            f"price_{var_previous}",
            f"price_{var_current}",
            "price_gap_percent",
        ]
    ]

    wb = create_excel_base()
    ws = wb.create_sheet(title="Out House")

    current_row = 3  # Starting row for data

    # Group data by part_num to identify boundaries
    grouped_data = df.groupby('part_num', sort=False)
    group_keys = list(grouped_data.groups.keys())

    for group_idx, (part_num, group_df) in enumerate(grouped_data):
        # Add data rows for current group
        for _, row in group_df.iterrows():
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=c_idx, value=value)

                # Apply conditional formatting to Gap column (price_gap_percent)
                if c_idx == 7:  # Adjusted column index for price_gap_percent
                    apply_threshold_formatting(cell, value, bounderies)

            current_row += 1

        # Insert empty row after each group (except the last one)
        if group_idx < len(group_keys) - 1:
            # Create empty row with just borders/formatting if needed
            for c_idx in range(1, 8):  # Columns 1-7
                cell = ws.cell(row=current_row, column=c_idx, value="")
                cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

            current_row += 1

    # Create headers
    create_styled_header(ws, "5 Part No", 1, 1, 2, 1)
    create_styled_header(ws, "Part No", 1, 2, 2, 2)
    create_styled_header(ws, "Part Name", 1, 3, 2, 3)
    create_styled_header(ws, "Source", 1, 4, 2, 4)
    create_styled_header(ws, "Price", 1, 5, 1, 7)

    create_styled_header(ws, f"{var_previous}", 2, 5, 2, 5)
    create_styled_header(ws, f"{var_current}", 2, 6, 2, 6)
    create_styled_header(ws, "Gap (%)", 2, 7, 2, 7)

    # Apply borders - adjust the end row to account for added empty rows
    total_rows = current_row - 1
    apply_borders(ws, 1, 1, 1, 7)  # Header borders
    apply_borders(ws, 2, 1, total_rows, 7)  # Data borders

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def convert_to_excel_format_in_house_per_part(df, var_previous, var_current, bounderies):
    df = df[
        [
            "part_num",
            "part_no",
            "part_name",
            f"lva_{var_previous}",
            f"lva_{var_current}",
            f"non_lva_{var_previous}",
            f"non_lva_{var_current}",
            f"process_{var_previous}",
            f"process_{var_current}",
            f"tooling_{var_previous}",
            f"tooling_{var_current}",
            f"total_cost_{var_previous}",
            f"total_cost_{var_current}",
            "price_gap_percent",
        ]
    ]

    wb = create_excel_base()
    ws = wb.create_sheet(title="In House")

    current_row = 3  # Starting row for data

    # Group data by part_num to identify boundaries
    grouped_data = df.groupby('part_num', sort=False)
    group_keys = list(grouped_data.groups.keys())

    for group_idx, (part_num, group_df) in enumerate(grouped_data):
        # Add data rows for current group
        for _, row in group_df.iterrows():
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=c_idx, value=value)

                # Apply conditional formatting to Gap column (price_gap_percent)
                if c_idx == 14:  # Adjusted column index for price_gap_percent
                    apply_threshold_formatting(cell, value, bounderies)

            current_row += 1

        # Insert empty row after each group (except the last one)
        if group_idx < len(group_keys) - 1:
            # Create empty row with just borders/formatting if needed
            for c_idx in range(1, 15):  # Columns 1-7
                cell = ws.cell(row=current_row, column=c_idx, value="")
                cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

            current_row += 1

    # Create headers
    create_styled_header(ws, "Part No", 1, 1, 2, 1)  # No merge for single cell
    create_styled_header(ws, "Part No", 1, 2, 2, 2)  # No merge for single cell
    create_styled_header(ws, "Part Name", 1, 3, 2, 3)  # No merge for single cell
    create_styled_header(ws, "LVA", 1, 4, 1, 5)  # Merge across 2 columns
    create_styled_header(ws, "Non LVA", 1, 6, 1, 7)  # Merge across 2 columns
    create_styled_header(ws, "Process Cost", 1, 8, 1, 9)  # Merge across 2 columns
    create_styled_header(ws, "Tooling Cost", 1, 10, 1, 11)  # Merge across 2 columns
    create_styled_header(ws, "Total Cost", 1, 12, 1, 13)  # Merge across 2 columns
    create_styled_header(ws, "Gap (%)", 1, 14, 2, 14)  # No merge for single cell

    # Second row headers (sub-categories) - only for merged columns
    create_styled_header(ws, f"{var_previous}", 2, 4, 2, 4)
    create_styled_header(ws, f"{var_current}", 2, 5, 2, 5)
    create_styled_header(ws, f"{var_previous}", 2, 6, 2, 6)
    create_styled_header(ws, f"{var_current}", 2, 7, 2, 7)
    create_styled_header(ws, f"{var_previous}", 2, 8, 2, 8)
    create_styled_header(ws, f"{var_current}", 2, 9, 2, 9)
    create_styled_header(ws, f"{var_previous}", 2, 10, 2, 10)
    create_styled_header(ws, f"{var_current}", 2, 11, 2, 11)
    create_styled_header(ws, f"{var_previous}", 2, 12, 2, 12)
    create_styled_header(ws, f"{var_current}", 2, 13, 2, 13)

    # Apply borders - adjust the end row to account for added empty rows
    total_rows = current_row - 1
    apply_borders(ws, 1, 1, 1, 14)  # Header borders
    apply_borders(ws, 2, 1, total_rows, 14)  # Data borders

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def convert_to_excel_format_packaging(df, var_previous, var_current, bounderies):
    # Filter columns
    df = df[
        [
            "part_no",
            "part_name",
            "destination",
            f"Labor Cost {var_previous}",
            f"Labor Cost {var_current}",
            f"Material Cost {var_previous}",
            f"Material Cost {var_current}",
            f"Inland Cost {var_previous}",
            f"Inland Cost {var_current}",
            f"Max Total Cost {var_previous}",
            f"Max Total Cost {var_current}",
            "Gap Total Cost",
        ]
    ]

    wb = create_excel_base()
    ws = wb.create_sheet(title=str("packing"))

    # Process each unique destination
    # for destination in df["destination"].unique():
    #     dest_df = df[df["destination"] == destination]

    # Populate data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Apply formatting to Gap Total Cost
            if c_idx == 12:
                apply_threshold_formatting(cell, value, bounderies)

        # Create headers
    create_styled_header(ws, "Part No", 1, 1, 2, 1)
    create_styled_header(ws, "Part Name", 1, 2, 2, 2)
    create_styled_header(ws, "Destination", 1, 3, 2, 3)

    # Create cost category headers
    cost_headers = ["Labor Cost", "Material Cost", "Inland Cost", "Total Cost"]
    for i, header in enumerate(cost_headers):
        col_start = 4 + i * 2
        create_styled_header(ws, header, 1, col_start, 1, col_start + 1)
        create_styled_header(ws, f"{var_previous}", 2, col_start, 2, col_start)
        create_styled_header(ws, f"{var_current}", 2, col_start + 1, 2, col_start + 1)

    create_styled_header(ws, "Gap Total Cost (%)", 1, 12, 2, 12)

    # Apply borders
    apply_borders(ws, 1, 1, 1, 12)
    apply_borders(ws, 2, 1, len(df) + 2, 12)


    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file
